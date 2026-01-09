import streamlit as st
import pdfplumber
import re
import requests
from openpyxl import load_workbook
from io import BytesIO

# =========================
# BRØNNØYSUND API
# =========================
def lookup_by_org_number(org_number):
    url = f"https://data.brreg.no/enhetsregisteret/api/enheter/{org_number}"
    r = requests.get(url, timeout=10)
    return r.json() if r.status_code == 200 else None


def search_company_by_name(name):
    url = "https://data.brreg.no/enhetsregisteret/api/enheter"
    params = {"navn": name}
    r = requests.get(url, params=params, timeout=10)
    if r.status_code == 200:
        data = r.json()
        if "_embedded" in data and data["_embedded"].get("enheter"):
            return data["_embedded"]["enheter"][0]
    return None


# =========================
# PDF TEXT EXTRACTION
# =========================
def extract_pdf_text(pdf_file):
    text = ""
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text


# =========================
# VERY LIGHT PDF FIELD EXTRACTION
# =========================
def extract_fields_from_text(text):
    fields = {}

    # Only org number is realistic to extract reliably
    org_match = re.search(r"\b\d{9}\b", text)
    fields["org_number"] = org_match.group(0) if org_match else ""

    return fields


# =========================
# EXCEL UPDATE
# =========================
def update_excel(template_file, data, summary):
    wb = load_workbook(template_file)
    ws = wb.active

    cell_mapping = {
        "company_name": "B14",
        "org_number": "B15",
        "address": "B16",
        "post_nr": "B17",
        "nace_code": "B18",
        "homepage": "B21",
        "employees": "B22",
    }

    for field, cell in cell_mapping.items():
        if data.get(field):
            ws[cell] = data[field]

    if summary:
        ws["B10"] = f"Kort info om företaget:\n{summary}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =========================
# STREAMLIT UI
# =========================
st.set_page_config(page_title="PDF → Excel (Brreg)", layout="centered")
st.title("📄➡️📊 PDF → Excel (Brønnøysund)")

pdf_file = st.file_uploader("Upload PDF (optional)", type="pdf")
excel_file = st.file_uploader("Upload Excel template", type="xlsx")

manual_company_name = st.text_input(
    "Company name (recommended)",
    placeholder="e.g. Eksempel AS"
)

if excel_file and st.button("Extract & Update Excel"):
    with st.spinner("Processing…"):

        extracted = {}

        # STEP 1 — PDF (optional)
        if pdf_file:
            pdf_text = extract_pdf_text(pdf_file)
            extracted = extract_fields_from_text(pdf_text)

        # STEP 2 — Determine lookup keys
        company_name = manual_company_name.strip()
        org_number = extracted.get("org_number", "")

        # STEP 3 — Brønnøysund lookup
        company_data = None
        if org_number:
            company_data = lookup_by_org_number(org_number)

        if not company_data and company_name:
            company_data = search_company_by_name(company_name)

        # STEP 4 — Normalize data
        if company_data:
            extracted["company_name"] = company_data.get("navn", "")
            extracted["org_number"] = company_data.get("organisasjonsnummer", "")

            addr = company_data.get("forretningsadresse") or {}
            extracted["address"] = " ".join(addr.get("adresse", []))
            extracted["post_nr"] = addr.get("postnummer", "")

            nace = company_data.get("naeringskode1", {})
            extracted["nace_code"] = nace.get("kode", "")

            extracted["homepage"] = company_data.get("hjemmeside", "")
            extracted["employees"] = company_data.get("antallAnsatte", "")

            summary = extracted["company_name"]
            if nace.get("beskrivelse"):
                summary += f" – {nace['beskrivelse']}"
        else:
            summary = ""

        updated_excel = update_excel(excel_file, extracted, summary)

    st.success("Excel updated successfully")
    st.json(extracted)

    st.download_button(
        "Download updated Excel",
        data=updated_excel,
        file_name="updated_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
