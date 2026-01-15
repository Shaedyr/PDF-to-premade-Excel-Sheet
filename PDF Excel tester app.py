import os
import re
import difflib
import requests
import wikipedia
import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pdfplumber

# =========================
# CONFIGURATION
# =========================
st.set_page_config(page_title="PDF → Excel (Brønnøysund)", layout="wide", page_icon="📊")
for k, v in {"extracted_data": {}, "api_response": None, "excel_ready": False, "company_summary": ""}.items():
    if k not in st.session_state:
        st.session_state[k] = v
for k in ('selected_company_data', 'companies_list', 'current_search', 'last_search', 'show_dropdown'):
    if k not in st.session_state:
        st.session_state[k] = None if k == 'selected_company_data' else [] if k == 'companies_list' else "" if k in ('current_search','last_search') else False

# =========================
# HELPERS: WIKI / WEB / SUMMARY
# =========================
def _strip_suffix(name: str):
    return re.sub(r'\b(AS|ASA|ANS|DA|ENK|KS|BA)\b\.?$', '', (name or ''), flags=re.I).strip()

def _wiki_summary(name: str, prefer_name: str = None):
    if not name: return None
    base = _strip_suffix(name)
    attempts = [name, base, base + " (bedrift)", base + " (company)"]
    candidates = []
    tried = set()
    for a in attempts:
        try:
            for r in wikipedia.search(a)[:8]:
                if r not in tried:
                    tried.add(r); candidates.append(r)
        except Exception:
            continue
    def score(title, summary):
        t, s = (title or "").lower(), (summary or "").lower()
        sc = 0
        pref = (prefer_name or base).lower()
        if pref and pref in t: sc += 50
        if pref and pref in s: sc += 20
        for kw in ("as","asa","bedrift","selskap","company","firma","group"):
            if kw in t or kw in s: sc += 10
        if len(s) > 80: sc += 2
        return sc
    best, best_s = None, -1
    for cand in candidates:
        try:
            page = wikipedia.page(cand, auto_suggest=False)
            s = page.summary or ""
            sc = score(page.title, s)
            if sc >= 60:
                sent = [x.strip() for x in s.split('. ') if x.strip()]
                return '. '.join(sent[:2]) + '.' if len(sent) > 2 else (s[:300] + '...' if len(s) > 300 else s)
            if sc > best_s and s:
                best_s, best = sc, s
        except Exception:
            continue
    if best and best_s >= 15:
        sent = [x.strip() for x in best.split('. ') if x.strip()]
        return '. '.join(sent[:2]) + '.' if len(sent) > 2 else (best[:300] + '...' if len(best) > 300 else best)
    return None

def _web_summary(q: str):
    if not q: return None
    q = _strip_suffix(q)
    for term in ("bedrift","company"):
        try:
            url = f"https://api.duckduckgo.com/?q={requests.utils.quote(q+' '+term)}&format=json&no_html=1&skip_disambig=1"
            r = requests.get(url, timeout=10)
            if r.status_code == 200:
                txt = r.json().get("AbstractText","") or ""
                if len(txt) > 50:
                    return txt if term=="bedrift" else txt.replace(" is a "," er et ").replace(" company"," selskap").replace(" based in "," med hovedkontor i ")
        except Exception:
            continue
    return None

def create_summary_from_brreg_data(d: dict):
    name = d.get("company_name","")
    if not name: return "Ingen informasjon tilgjengelig om dette selskapet."
    parts = []
    industry, city, emp, reg = d.get("nace_description",""), d.get("city",""), d.get("employees",""), d.get("registration_date","")
    if industry and city: parts.append(f"{name} driver {industry.lower()} virksomhet fra {city}.")
    elif industry: parts.append(f"{name} opererer innen {industry.lower()}.")
    else: parts.append(f"{name} er et registrert norsk selskap.")
    if reg:
        try:
            year = int(reg.split('-')[0]) if '-' in reg else int(reg)
            age = datetime.now().year - year
            parts.append(f"Etablert i {year}, {age} år erfaring.") if year else parts.append(f"Registrert: {reg}.")
        except Exception:
            parts.append(f"Selskapet ble registrert i {reg}.")
    if emp:
        try:
            e = int(emp)
            if e > 200: parts.append(f"Større arbeidsgiver med {e} ansatte.")
            elif e > 50: parts.append(f"Mellomstort foretak med {e} ansatte.")
            elif e > 10: parts.append(f"Selskapet sysselsetter {e} personer.")
            else: parts.append(f"Lite selskap med {e} ansatte.")
        except Exception:
            pass
    if len(parts) < 3: parts.append("Virksomheten er registrert i Brønnøysundregistrene.")
    s = ' '.join(parts)
    return s[:797] + "..." if len(s) > 800 else s

# =========================
# BRØNNØYSUND SEARCH
# =========================
@st.cache_data(ttl=3600)
def search_companies_live(name: str):
    if not name or len(name.strip()) < 2: return []
    try:
        r = requests.get("https://data.brreg.no/enhetsregisteret/api/enheter", params={"navn": name.strip(), "size": 10}, timeout=30)
        if r.status_code == 200:
            return r.json().get("_embedded", {}).get("enheter", []) or []
    except Exception:
        pass
    return []

def format_brreg_data(api_data):
    if not api_data: return {}
    out = {"company_name": api_data.get("navn",""), "org_number": api_data.get("organisasjonsnummer",""),
           "nace_code": "", "nace_description": "", "homepage": api_data.get("hjemmeside",""),
           "employees": api_data.get("antallAnsatte",""), "address": "", "post_nr": "", "city": "", "registration_date": api_data.get("stiftelsesdato","")}
    addr = api_data.get("forretningsadresse",{}) or {}
    if addr:
        a = addr.get("adresse",[])
        out["address"] = ", ".join(filter(None, a)) if isinstance(a,list) else (a or "")
        out["post_nr"] = addr.get("postnummer","")
        out["city"] = addr.get("poststed","")
    nace = api_data.get("naeringskode1",{}) or {}
    if nace:
        out["nace_code"] = nace.get("kode",""); out["nace_description"] = nace.get("beskrivelse","")
    return out

# =========================
# EXCEL: TEMPLATE LOADING AND FILLING
# =========================
TARGET_FILL_HEX = "F2F2F2"

def load_template_from_url_or_file(template_url=None, uploaded_file=None):
    """Load Excel template from URL or uploaded file"""
    try:
        # Priority 1: Use uploaded file if provided
        if uploaded_file:
            template_bytes = uploaded_file.read()
            # Verify it's a valid Excel file
            try:
                load_workbook(BytesIO(template_bytes))
                return template_bytes
            except Exception as e:
                st.error(f"Opplastet fil er ikke en gyldig Excel-fil: {e}")
                return None
        
        # Priority 2: Try to download from URL
        if template_url:
            # For OneDrive links, we need to use a different approach
            if "1drv.ms" in template_url:
                # OneDrive links are tricky - they often require authentication
                # Let's try a direct approach with requests
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                }
                
                # Try to get the file
                response = requests.get(template_url, headers=headers, timeout=30, allow_redirects=True)
                
                if response.status_code == 200:
                    content = response.content
                    # Check if it's an Excel file (starts with PK for zip files)
                    if content[:2] == b'PK':
                        return content
                    else:
                        # Might be HTML page, try to extract download link
                        html_content = response.text
                        # Look for direct download links in the HTML
                        import re
                        # Try to find direct download links
                        file_links = re.findall(r'href="([^"]+\.xlsx)"', html_content)
                        for link in file_links:
                            if 'download' in link.lower():
                                response2 = requests.get(link, headers=headers, timeout=30)
                                if response2.status_code == 200 and response2.content[:2] == b'PK':
                                    return response2.content
            
            # If URL approach fails, use GitHub as fallback
            fallback_url = "https://raw.githubusercontent.com/Shaedyr/PDF-to-premade-Excel-Sheet/main/PremadeExcelTemplate.xlsx"
            response = requests.get(fallback_url, timeout=30)
            if response.status_code == 200:
                return response.content
        
        # Priority 3: Try local file
        if os.path.exists("Grundmall.xlsx"):
            with open("Grundmall.xlsx", "rb") as f:
                return f.read()
        
        return None
        
    except Exception as e:
        st.error(f"Feil ved lasting av mal: {e}")
        return None

def _rgb_hex_from_color(col):
    if not col: return None
    rgb = getattr(col,"rgb",None)
    if not rgb: return None
    rgb = rgb.upper()
    if len(rgb)==8: rgb = rgb[2:]
    return rgb if len(rgb)==6 else None

FIELD_KEYWORDS = {
    "company_name": ["selskapsnavn","selskap","navn","firmanavn","firma","kunde","dagens selskap"],
    "org_number": ["organisasjonsnummer","orgnr","org.nr","org nr","orgnummer"],
    "address": ["adresse","gate","gatenavn","postadresse","adr"],
    "post_nr": ["postnummer","postnr","post nr","postkode"],
    "city": ["poststed","sted","by"],
    "employees": ["ansatte","antall ansatte","antal ansatte"],
    "homepage": ["hjemmeside","nettside","web","website","url"],
    "nace_code": ["nacekode","nace kode","nace","naeringskode"],
    "nace_description": ["bransje","næring"],
    "company_summary": ["om bedriften","sammendrag","om oss","om selskapet"],
    "revenue_2024": ["omsetning 2024","omsetning"],
    "tender_deadline": ["anbudsfrist","frist"]
}

def _normalize_label(t):
    return re.sub(r'[^a-zA-Z0-9æøåÆØÅ]+',' ', (t or "").lower()).strip()

def _match_field_by_label(label):
    if not label: return None
    lab = _normalize_label(label)
    for field, kws in FIELD_KEYWORDS.items():
        for kw in kws:
            if kw in lab: return field
    tokens = lab.split()
    for field, kws in FIELD_KEYWORDS.items():
        for kw in kws:
            if any(tok in tokens for tok in kw.split()): return field
    best_field, best_score = None, 0.0
    for field, kws in FIELD_KEYWORDS.items():
        for kw in kws:
            sc = difflib.SequenceMatcher(None, lab, kw).ratio()
            if sc > best_score:
                best_score, best_field = sc, field
    return best_field if best_score >= 0.60 else None

def scan_and_map_fill_cells(wb_bytes):
    mapping = {}; unmatched = []; debug = []
    wb = load_workbook(BytesIO(wb_bytes), data_only=False)
    for ws in wb.worksheets:
        fillables = []
        assigned = {}
        for row in ws.iter_rows():
            for c in row:
                try:
                    fg = getattr(c.fill,"fgColor",None) or getattr(c.fill,"start_color",None)
                    hexcol = _rgb_hex_from_color(fg)
                    is_fill = True if (hexcol and hexcol.upper()==TARGET_FILL_HEX) else False
                    label = ""
                    if c.column>1:
                        left = ws.cell(row=c.row, column=c.column-1).value
                        if left: label = str(left)
                    if not label and c.row>1:
                        above = ws.cell(row=c.row-1, column=c.column).value
                        if above: label = str(above)
                    debug.append((ws.title, c.coordinate, hexcol, is_fill, label))
                    if is_fill:
                        fillables.append((c.row, c.column, c.coordinate))
                except Exception:
                    continue
        for r,c,coord in fillables:
            neighbors = [(r,c-1),(r-1,c),(r,c+1),(r+1,c),(r-1,c-1),(r-1,c+1),(r+1,c-1),(r+1,c+1)]
            for rr,cc in neighbors:
                if rr<1 or cc<1: continue
                try:
                    val = ws.cell(row=rr, column=cc).value
                    if val:
                        f = _match_field_by_label(str(val))
                        if f and f not in assigned:
                            assigned[f] = (coord,"local",str(val)); break
                except Exception:
                    continue
        label_cells = []
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v and str(v).strip(): label_cells.append((c.row,c.column,c.coordinate,str(v)))
        all_fields = list(FIELD_KEYWORDS.keys())
        unassigned = [f for f in all_fields if f not in assigned]
        available = [coord for (_r,_c,coord) in fillables]
        def man(r1,c1,r2,c2): return abs(r1-r2)+abs(c1-c2)
        for field in unassigned:
            best_label=None; best_score=0.0
            for rr,cc,coord_label,raw in label_cells:
                lab = _normalize_label(raw)
                for kw in FIELD_KEYWORDS[field]:
                    if kw in lab:
                        best_label=(rr,cc,coord_label,raw); best_score=1.0; break
                if best_label is None:
                    for kw in FIELD_KEYWORDS[field]:
                        sc = difflib.SequenceMatcher(None, lab, kw).ratio()
                        if sc>best_score:
                            best_score=sc; best_label=(rr,cc,coord_label,raw)
            if best_label and best_score>=0.55:
                rr,cc,_,raw = best_label
                best_fill=None; best_dist=None
                for (_r,_c,coord_fill) in fillables:
                    if coord_fill not in available: continue
                    d = man(rr,cc,_r,_c)
                    if best_dist is None or d<best_dist:
                        best_dist, best_fill = d, coord_fill
                if best_fill:
                    assigned[field] = (best_fill,"global",raw)
                    if best_fill in available: available.remove(best_fill)
        smap = {f:v for f,v in assigned.items()}
        remaining_unmapped = [coord for (_r,_c,coord) in fillables if coord not in [v[0] for v in assigned.values()]]
        for coord_un in remaining_unmapped:
            lab = ""
            rcell = ws[coord_un].row; ccell = ws[coord_un].column
            if ccell>1:
                left = ws.cell(row=rcell, column=ccell-1).value
                if left: lab = str(left)
            if not lab and rcell>1:
                above = ws.cell(row=rcell-1, column=ccell).value
                if above: lab = str(above)
            unmatched.append((ws.title, coord_un, lab or None))
        if smap: mapping[ws.title] = smap
    return mapping, unmatched, debug

def fill_workbook_bytes(template_bytes: bytes, field_values: dict):
    report = {"filled": [], "skipped": [], "errors": [], "unmapped_cells": [], "debug_cells": [], "mapping": {}}
    wb_read = load_workbook(BytesIO(template_bytes), data_only=False)
    first_sheet = wb_read.sheetnames[0] if wb_read.sheetnames else None
    mapping, unmatched, debug = scan_and_map_fill_cells(template_bytes)
    report["debug_cells"] = debug; report["mapping"] = mapping
    first_map = mapping.get(first_sheet, {}) if first_sheet else {}
    wb = load_workbook(BytesIO(template_bytes))
    if not first_sheet:
        report["errors"].append(("NO_SHEET",None,"No sheets")); return template_bytes, report
    ws = wb[first_sheet]
    for field, val in field_values.items():
        if field in first_map:
            coord = first_map[field][0]
            try:
                if val not in (None,""):
                    ws[coord].value = str(val); report["filled"].append((first_sheet,coord,field))
                else:
                    report["skipped"].append((first_sheet,coord,field,"No value"))
            except Exception as e:
                report["errors"].append((first_sheet,coord,field,str(e)))
        else:
            report["skipped"].append((first_sheet,None,field,"No mapped cell on first sheet"))
    remaining = {k:v for k,v in field_values.items() if v not in (None,"")}
    for (_s,_coord,f) in report["filled"]:
        remaining.pop(f, None)
    unmatched_first = [t for t in unmatched if t[0]==first_sheet]
    if unmatched_first and remaining:
        for (sheetname, coord, label) in unmatched_first:
            if not remaining: break
            field_name, val = remaining.popitem()
            try:
                wb[sheetname][coord].value = str(val); report["filled"].append((sheetname,coord,field_name,"auto-mapped"))
            except Exception as e:
                report["errors"].append((sheetname,coord,field_name,str(e)))
    summary = field_values.get("company_summary") or ""
    if summary:
        wrote=False
        try:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for c in row:
                    if c.value and isinstance(c.value,str) and "skriv her" in c.value.strip().lower():
                        c.value = str(summary); c.alignment = Alignment(wrap_text=True, vertical='top')
                        report["filled"].append((first_sheet,c.coordinate,"company_summary","replaced 'Skriv her'")); wrote=True; break
                if wrote: break
            if not wrote:
                ws["A46"] = str(summary); ws["A46"].alignment = Alignment(wrap_text=True, vertical='top')
                report["filled"].append((first_sheet,"A46","company_summary","fallback A46"))
        except Exception as e:
            report["errors"].append((first_sheet,None,"company_summary",str(e)))
    out = BytesIO(); wb.save(out); out.seek(0)
    report["unmapped_cells"] = [u for u in unmatched if u[0]==first_sheet]
    return out.getvalue(), report

# =========================
# PDF EXTRACTION
# =========================
ORG_RE = re.compile(r'\b(\d{9})\b')
ORG_IN_TEXT_RE = re.compile(r'(organisasjonsnummer|org\.?nr|org nr|orgnummer)[:\s]*?(\d{9})', flags=re.I)
COMPANY_WITH_SUFFIX_RE = re.compile(r'([A-ZÆØÅ][A-Za-zÆØÅæøå\.\-&\s]{1,120}?)\s+(AS|ASA|ANS|DA|ENK|KS|BA)\b', flags=re.I)

def extract_text_from_pdf(file_bytes):
    try:
        text = ""
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            for p in pdf.pages[:6]:
                text += p.extract_text() or ""
        return text
    except Exception:
        return ""

def extract_fields_from_pdf_bytes(pdf_bytes):
    txt = extract_text_from_pdf(pdf_bytes); fields = {}
    m = ORG_IN_TEXT_RE.search(txt)
    if m: fields["org_number"] = m.group(2)
    else:
        m2 = ORG_RE.search(txt)
        if m2: fields["org_number"] = m2.group(1)
    m3 = COMPANY_WITH_SUFFIX_RE.search(txt)
    if m3: fields["company_name"] = m3.group(0).strip()
    else:
        lines = [l.strip() for l in txt.splitlines() if l.strip()]
        for ln in lines[:30]:
            if len(ln)>3 and any(ch.isalpha() for ch in ln) and ln==ln.title():
                fields["company_name"] = ln; break
    mpc = re.search(r'(\d{4})\s+([A-ZÆØÅa-zæøå\-\s]{2,50})', txt)
    if mpc: fields["post_nr"] = mpc.group(1); fields["city"] = mpc.group(2).strip()
    maddr = re.search(r'([A-ZÆØÅa-zæøå\.\-\s]{3,60}\s+\d{1,4}[A-Za-z]?)', txt)
    if maddr: fields["address"] = maddr.group(1).strip()
    mrev = re.search(r'omsetning\s*(?:2024)?[:\s]*([\d\s\.,]+(?:kr)?)', txt, flags=re.I)
    if mrev: fields["revenue_2024"] = mrev.group(1).strip()
    mdate = re.search(r'(?:anbudsfrist|frist)[:\s]*([0-3]?\d[./-][01]?\d[./-]\d{2,4})', txt, flags=re.I)
    if mdate: fields["tender_deadline"] = mdate.group(1).strip()
    return fields

def fetch_brreg_by_org(org_number: str):
    try:
        r = requests.get(f"https://data.brreg.no/enhetsregisteret/api/enheter/{org_number}", timeout=20)
        if r.status_code==200: return r.json()
    except Exception:
        pass
    return None

# =========================
# STREAMLIT UI
# =========================
def main():
    st.title("📄 PDF → Excel (Brønnøysund)")
    st.markdown("Hent selskapsinformasjon og oppdater Excel automatisk")
    st.markdown("---")
    
    # OneDrive template URL
    ONEDRIVE_TEMPLATE_URL = "https://1drv.ms/x/c/f5e2800feeb07258/IQBBPI2scMXjQ6bi18LIvXFGAWFnYqG3J_kCKfewCEid9Bc"
    
    st.markdown("### 📋 Excel-mal")
    
    # Upload template section
    uploaded_template = st.file_uploader(
        "Last opp Excel-mal (anbefalt)",
        type=["xlsx"],
        help="Last opp Excel-malen din direkte for å sikre at den fungerer",
        key="template_upload"
    )
    
    if uploaded_template:
        st.success("✅ Excel-mal lastet opp!")
    
    # OR use OneDrive URL
    st.markdown("---")
    st.markdown("**ELLER** bruk OneDrive lenke:")
    
    col_url1, col_url2 = st.columns([3, 1])
    
    with col_url1:
        template_url = st.text_input(
            "OneDrive lenke (valgfritt):",
            value=ONEDRIVE_TEMPLATE_URL,
            help="Lim inn OneDrive lenken din",
            key="template_url_input"
        )
    
    # Main content columns
    c1, c2 = st.columns(2)
    
    with c1:
        pdf_file = st.file_uploader("PDF dokument (valgfritt)", type="pdf", help="Last opp PDF for referanse")
    
    with c2:
        q = st.text_input("Selskapsnavn *", placeholder="Skriv her... (minst 2 bokstaver)", key="company_search_input")
        if st.session_state.get('current_search','') != q:
            st.session_state.selected_company_data = None
        st.session_state.current_search = q
        if q and len(q.strip())>=2:
            with st.spinner("Søker..."):
                comps = search_companies_live(q)
            if comps:
                opts=["-- Velg selskap --"]; cd={}
                for c in comps:
                    name = c.get('navn','Ukjent'); org=c.get('organisasjonsnummer',''); city=c.get('forretningsadresse',{}).get('poststed','')
                    disp = f"{name}" + (f" (Org.nr: {org})" if org else "") + (f" - {city}" if city else "")
                    opts.append(disp); cd[disp]=c
                sel = st.selectbox("🔍 Søkeresultater:", opts, key="dynamic_company_dropdown")
                if sel and sel!="-- Velg selskap --":
                    st.session_state.selected_company_data = format_brreg_data(cd[sel]); st.success(f"✅ Valgt: {cd[sel].get('navn')}")
                else:
                    if len(q.strip())>=3: st.warning("Ingen selskaper funnet. Prøv et annet navn.")
                    st.session_state.selected_company_data = None

    st.markdown("---")
    
    # Load Excel template when needed
    if 'template_loaded' not in st.session_state:
        # Try to load template
        template_bytes = None
        
        # Priority 1: Use uploaded file
        if uploaded_template:
            try:
                uploaded_template.seek(0)
                template_bytes = uploaded_template.read()
                load_workbook(BytesIO(template_bytes))  # Verify it's valid
                st.session_state.template_bytes = template_bytes
                st.session_state.template_loaded = True
                st.success("✅ Excel-mal lastet fra opplastet fil!")
            except Exception as e:
                st.error(f"❌ Feil ved lasting av opplastet mal: {e}")
        
        # Priority 2: Try OneDrive URL
        elif not st.session_state.get('template_loaded') and template_url:
            with st.spinner("Prøver å laste fra OneDrive..."):
                template_bytes = load_template_from_url_or_file(template_url=template_url)
                if template_bytes:
                    st.session_state.template_bytes = template_bytes
                    st.session_state.template_loaded = True
                    st.success("✅ Excel-mal lastet fra OneDrive!")
                else:
                    st.error("❌ Kunne ikke laste fra OneDrive. Last opp filen direkte.")
        
        # Priority 3: Use fallback
        if not st.session_state.get('template_loaded'):
            with st.spinner("Bruker standard mal..."):
                template_bytes = load_template_from_url_or_file()
                if template_bytes:
                    st.session_state.template_bytes = template_bytes
                    st.session_state.template_loaded = True
                    st.success("✅ Standard Excel-mal lastet!")
                else:
                    st.session_state.template_loaded = False
                    st.error("❌ Ingen Excel-mal tilgjengelig. Last opp en mal.")
    
    # Reload template if new file uploaded
    elif uploaded_template and uploaded_template != st.session_state.get('last_uploaded_template'):
        try:
            uploaded_template.seek(0)
            template_bytes = uploaded_template.read()
            load_workbook(BytesIO(template_bytes))  # Verify it's valid
            st.session_state.template_bytes = template_bytes
            st.session_state.last_uploaded_template = uploaded_template
            st.success("✅ Ny Excel-mal lastet!")
        except Exception as e:
            st.error(f"❌ Feil ved lasting av ny mal: {e}")
    
    elif st.session_state.get('template_loaded'):
        st.info("✅ Excel-mal er klar for bruk.")

    st.markdown("---")
    
    # Process button
    if st.button("🚀 Prosesser & Oppdater Excel", use_container_width=True, type="primary"):
        if not st.session_state.get('template_loaded'): 
            st.error("❌ Excel-mal ikke tilgjengelig")
            st.stop()
        
        field_values = {}
        if st.session_state.selected_company_data:
            field_values.update(st.session_state.selected_company_data)
        
        if pdf_file:
            try:
                pdf_bytes = pdf_file.read()
                extracted = extract_fields_from_pdf_bytes(pdf_bytes)
                if "org_number" in extracted:
                    br = fetch_brreg_by_org(extracted["org_number"])
                    if br:
                        br_data = format_brreg_data(br)
                        for k,v in br_data.items():
                            if v: field_values[k]=v
                        for k,v in extracted.items():
                            if v: field_values[k]=v
                    else:
                        field_values.update(extracted)
                else:
                    field_values.update(extracted)
            except Exception as e:
                st.error(f"❌ Feil ved PDF-parsing: {e}")

        if not field_values:
            st.error("❌ Ingen selskapsdata funnet.")
            st.stop()

        # Get company summary
        company_summary = None
        brreg_like = st.session_state.get("selected_company_data") or {}
        if brreg_like:
            company_summary = create_summary_from_brreg_data(brreg_like)
        if not company_summary or (isinstance(company_summary,str) and len(company_summary)<40):
            name = field_values.get("company_name","") or ""
            if name:
                try:
                    company_summary = _wiki_summary(name, prefer_name=name)
                except Exception:
                    company_summary = None
        if not company_summary or (isinstance(company_summary,str) and len(company_summary)<40):
            try:
                company_summary = _web_summary(field_values.get("company_name","") or field_values.get("org_number",""))
            except Exception:
                company_summary = None
        if not company_summary:
            company_summary = create_summary_from_brreg_data(field_values)
        
        field_values["company_summary"] = company_summary or ""
        st.session_state.company_summary = company_summary or ""
        st.session_state.extracted_data = field_values

        try:
            updated_bytes, report = fill_workbook_bytes(st.session_state.template_bytes, field_values)
            st.session_state.excel_bytes = updated_bytes
            
            if report["filled"]: 
                st.success(f"✅ Fylte {len(report['filled'])} celler i Excel-malen.")
            else: 
                st.warning("Kunne ikke fylle noen celler — sjekk malstrukturen.")
            
            if report["errors"]:
                st.error("Noen celler kunne ikke fylles:")
                for err in report["errors"][:3]:
                    st.write(f"- {err}")
            
            st.session_state.excel_ready = True
            
        except Exception as e:
            st.error(f"❌ Feil ved utfylling av Excel: {e}")
            st.session_state.excel_ready = False

    # Display extracted data
    if st.session_state.extracted_data:
        st.markdown("---")
        st.subheader("📋 Ekstraherte data")
        d1,d2 = st.columns(2)
        with d1:
            d = st.session_state.extracted_data
            st.write(f"**Selskapsnavn:** {d.get('company_name','')}")
            st.write(f"**Organisasjonsnummer:** {d.get('org_number','')}")
            st.write(f"**Adresse:** {d.get('address','')}")
            st.write(f"**Postnummer:** {d.get('post_nr','')}")
            st.write(f"**Poststed:** {d.get('city','')}")
            st.write(f"**Antall ansatte:** {d.get('employees','')}")
            st.write(f"**Hjemmeside:** {d.get('homepage','')}")
            nc, nd = d.get('nace_code',''), d.get('nace_description','')
            if nc and nd: 
                st.write(f"**NACE-bransje/nummer:** {nd} ({nc})")
            elif nc: 
                st.write(f"**NACE-nummer:** {nc}")
            elif nd: 
                st.write(f"**NACE-bransje:** {nd}")
        
        with d2:
            if st.session_state.company_summary:
                st.write("**Sammendrag (går i 'Om oss' / 'Skriv her' celle):**")
                st.info(st.session_state.company_summary)

    # Download
    if st.session_state.get('excel_ready') and st.session_state.get('excel_bytes'):
        st.markdown("---")
        st.subheader("📥 Last ned ferdig Excel-fil")
        cname = st.session_state.extracted_data.get('company_name','selskap')
        safe = re.sub(r'[^\w\s-]','', cname, flags=re.UNICODE)
        safe = re.sub(r'[-\s]+','_', safe)
        
        st.download_button(
            label="⬇️ Last ned oppdatert Excel",
            data=st.session_state.excel_bytes,
            file_name=f"{safe}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
    
    st.markdown("---")
    st.caption("Drevet av Brønnøysund Enhetsregisteret API")

if __name__ == "__main__":
    main()
